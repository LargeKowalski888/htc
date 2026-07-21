from __future__ import annotations

import argparse
import base64
import json
from pathlib import Path
from datetime import datetime
from typing import Any

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import yaml


def find_prediction_table(run_dir: Path) -> Path:
    candidates = sorted(run_dir.rglob("test_predictions.csv"), key=lambda p: p.stat().st_mtime, reverse=True)
    if not candidates:
        raise FileNotFoundError(f"No test_predictions.csv found under {run_dir}")
    return candidates[0]


def flatten_dict(d: dict[str, Any], prefix: str = "") -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for k, v in d.items():
        key = f"{prefix}.{k}" if prefix else str(k)
        if isinstance(v, dict):
            rows.extend(flatten_dict(v, key))
        elif isinstance(v, list):
            rows.append({"setting": key, "value": json.dumps(v, ensure_ascii=False)})
        else:
            rows.append({"setting": key, "value": str(v)})
    return rows


def safe_load_json(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}


def row_keys(value: str) -> set[str]:
    value = str(value)
    base = value.split("@")[0]
    ts = base.split("#")[-1]
    return {value, base, ts}


def build_split_assignment(output_dir: Path, labels: list[str]) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, dict[str, Any]]:
    """Return individual/recording split summaries and long split details.

    Definitions used in the report:
      individual = subject_name, e.g. P001_OP001_2023_07_28_Experiment1
      recording  = timestamp/image, e.g. 2023_07_28_20_39_26

    The splitter itself is subject-based in the launcher, so one individual should
    not appear in multiple splits. This function checks and reports violations.
    """
    manifest_path = output_dir / "median_pixel_yaml_manifest.csv"
    data_spec_path = output_dir / "yaml_generated_median_pixel_data_spec_htc_adapter_annotated.json"
    info: dict[str, Any] = {"manifest_path": str(manifest_path), "data_spec_path": str(data_spec_path)}

    empty_summary = pd.DataFrame(columns=["label", "train", "val", "test", "total"])
    empty_details = pd.DataFrame(columns=["split", "individual", "recording", "label_name", "image_name", "sample_dir", "hypergui_dir", "spectrum_xlsx_path"])

    if not manifest_path.exists() or not data_spec_path.exists():
        info["warning"] = "manifest or data_spec missing"
        return empty_summary, empty_summary, empty_details, info

    manifest = pd.read_csv(manifest_path)
    spec = json.loads(data_spec_path.read_text(encoding="utf-8"))[0]

    required = {"subject_name", "timestamp", "image_name", "label_name"}
    missing = sorted(required - set(manifest.columns))
    if missing:
        info["warning"] = f"required manifest columns missing: {missing}"
        return empty_summary, empty_summary, empty_details, info

    # Build split keys from data spec. Matching accepts full name, name without @annotation, and timestamp only.
    split_keysets: dict[str, set[str]] = {}
    for split in ["train", "val", "test"]:
        keyset: set[str] = set()
        for name in spec.get(split, {}).get("image_names", []):
            keyset |= row_keys(str(name))
        split_keysets[split] = keyset

    details_rows: list[dict[str, Any]] = []
    for _, row in manifest.iterrows():
        keys = row_keys(str(row["image_name"])) | row_keys(str(row["timestamp"]))
        assigned = [split for split, keyset in split_keysets.items() if bool(keys & keyset)]
        if not assigned:
            continue
        split = assigned[0]
        details_rows.append({
            "split": split,
            "individual": str(row["subject_name"]),
            "recording": str(row["timestamp"]),
            "label_name": str(row["label_name"]),
            "image_name": str(row.get("image_name", "")),
            "sample_dir": str(row.get("sample_dir", "")),
            "hypergui_dir": str(row.get("hypergui_dir", "")),
            "spectrum_xlsx_path": str(row.get("spectrum_xlsx_path", "")),
        })

    details = pd.DataFrame(details_rows)
    if details.empty:
        info["warning"] = "no manifest rows matched data spec split names"
        return empty_summary, empty_summary, empty_details, info

    # Individual leakage check: one individual must belong to only one split.
    indiv_splits = details.groupby("individual")["split"].nunique()
    leaking = indiv_splits[indiv_splits > 1].index.tolist()
    info["individual_leakage_count"] = len(leaking)
    info["individual_leakage_examples"] = leaking[:10]

    def summary_unique(unique_col: str) -> pd.DataFrame:
        rows: list[dict[str, Any]] = []
        for label in labels:
            out = {"label": label}
            for split in ["train", "val", "test"]:
                d = details[(details["split"].eq(split)) & (details["label_name"].eq(label))]
                out[split] = int(d[unique_col].nunique())
            out["total"] = int(details[details["label_name"].eq(label)][unique_col].nunique())
            rows.append(out)
        return pd.DataFrame(rows)

    individual_counts = summary_unique("individual")
    recording_counts = summary_unique("recording")
    return individual_counts, recording_counts, details, info


def write_clean_excel_report(
    xlsx_path: Path,
    png_path: Path,
    labels: list[str],
    per_class: list[dict[str, Any]],
    mat_norm: np.ndarray,
    mat_counts: np.ndarray,
    individual_counts: pd.DataFrame,
    recording_counts: pd.DataFrame,
    split_details: pd.DataFrame,
    yaml_settings: dict[str, Any],
    run_summary: dict[str, Any],
    generated_configs: dict[str, Any],
    metadata: dict[str, Any],
) -> None:
    """Create a readable xlsx report with long paths isolated on their own sheet."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage

    wb = Workbook()
    ws = wb.active
    ws.title = "Overview"

    fill_title = PatternFill("solid", fgColor="1F4E78")
    fill_header = PatternFill("solid", fgColor="E2F0D9")
    fill_warn = PatternFill("solid", fgColor="FCE4D6")
    thin = Side(style="thin", color="D9D9D9")
    hair = Side(style="hair", color="E6E6E6")

    def write_df(sheet, df: pd.DataFrame, start_row: int = 1, start_col: int = 1, freeze: bool = True):
        if df is None or df.empty:
            return
        for j, col in enumerate(df.columns, start_col):
            c = sheet.cell(start_row, j, col)
            c.font = Font(bold=True)
            c.fill = fill_header
            c.border = Border(bottom=thin)
            c.alignment = Alignment(horizontal="center")
        for i, row in enumerate(df.itertuples(index=False), start_row + 1):
            for j, val in enumerate(row, start_col):
                c = sheet.cell(i, j, val)
                c.border = Border(bottom=hair)
        if freeze:
            sheet.freeze_panes = sheet.cell(start_row + 1, start_col + 1).coordinate
        sheet.auto_filter.ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(start_col + len(df.columns)-1)}{start_row + len(df)}"

    # Overview
    ws["A1"] = "Confusion Matrix Report"
    ws["A1"].fill = fill_title
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    ws.merge_cells("A1:H1")
    overview_rows = [
        ("Accuracy", metadata.get("accuracy", "")),
        ("Test rows", metadata.get("n_test_rows", "")),
        ("Profile", metadata.get("profile", "")),
        ("Created at", metadata.get("created_at", "")),
        ("Run folder", metadata.get("run_dir", "")),
        ("Prediction table", metadata.get("prediction_table", "")),
        ("YAML path", metadata.get("yaml_path", "")),
    ]
    for r, (k, v) in enumerate(overview_rows, 3):
        ws.cell(r, 1, k).font = Font(bold=True)
        ws.cell(r, 2, v)
        ws.cell(r, 2).alignment = Alignment(wrap_text=False)
    if png_path.exists():
        img = XLImage(str(png_path))
        img.width = 760
        img.height = 560
        ws.add_image(img, "D3")
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 70

    # Per-class metrics
    ws_metrics = wb.create_sheet("Per-Class Metrics")
    df_metrics = pd.DataFrame(per_class)
    write_df(ws_metrics, df_metrics)
    for col in range(1, len(df_metrics.columns) + 1):
        ws_metrics.column_dimensions[get_column_letter(col)].width = 18
    if "sensitivity" in df_metrics.columns:
        sens_col = list(df_metrics.columns).index("sensitivity") + 1
        for row in range(2, len(df_metrics) + 2):
            ws_metrics.cell(row, sens_col).number_format = "0.0%"
        ws_metrics.conditional_formatting.add(
            f"{get_column_letter(sens_col)}2:{get_column_letter(sens_col)}{len(df_metrics)+1}",
            ColorScaleRule(start_type="num", start_value=0, start_color="F8696B", mid_type="num", mid_value=0.5, mid_color="FFEB84", end_type="num", end_value=1, end_color="63BE7B"),
        )

    # Split summaries
    for sheet_name, df_counts, count_type in [
        ("Split Individuals", individual_counts, "individuals"),
        ("Split Recordings", recording_counts, "recordings"),
    ]:
        sh = wb.create_sheet(sheet_name)
        df = df_counts.copy()
        for part in ["train", "val", "test"]:
            if part in df.columns and "total" in df.columns:
                df[f"{part}_pct"] = (df[part] / df["total"]).where(df["total"] != 0, 0)
        write_df(sh, df)
        widths = {"A": 18, "B": 10, "C": 10, "D": 10, "E": 10, "F": 12, "G": 12, "H": 12}
        for col, width in widths.items():
            sh.column_dimensions[col].width = width
        for row in range(2, len(df) + 2):
            for col in range(6, min(8, len(df.columns)) + 1):
                sh.cell(row, col).number_format = "0.0%"
            # Flag rows where val or test is zero.
            val = sh.cell(row, 3).value if sh.max_column >= 3 else None
            test = sh.cell(row, 4).value if sh.max_column >= 4 else None
            if val == 0 or test == 0:
                for col in range(1, len(df.columns) + 1):
                    sh.cell(row, col).fill = fill_warn
        sh["J1"] = f"Counts shown here are unique {count_type} per label and split."
        sh["J1"].font = Font(italic=True, color="666666")

    # Split detail tabs: grouped layout like the supervisor example workbook.
    # Each individual is written once, followed by that individual's recordings.
    for split, sheet_name in [("train", "Train Details"), ("val", "Validation Details"), ("test", "Test Details")]:
        sh = wb.create_sheet(sheet_name)
        d = split_details[split_details["split"].eq(split)].copy() if not split_details.empty else pd.DataFrame()
        sh["A1"] = sheet_name
        sh["A1"].font = Font(bold=True, size=14, color="FFFFFF")
        sh["A1"].fill = fill_title
        sh.merge_cells("A1:G1")

        headers = ["individual / recording", "label_name", "image_name", "sample_dir", "hypergui_dir", "spectrum_xlsx_path"]
        for col, header in enumerate(headers, 1):
            c = sh.cell(3, col, header)
            c.font = Font(bold=True)
            c.fill = fill_header
            c.border = Border(bottom=thin)
            c.alignment = Alignment(horizontal="center")

        row_i = 4
        if not d.empty:
            d = d[["individual", "recording", "label_name", "image_name", "sample_dir", "hypergui_dir", "spectrum_xlsx_path"]].sort_values(["individual", "recording", "label_name"])
            for individual, group in d.groupby("individual", sort=True):
                # Individual row appears once.
                sh.cell(row_i, 1, str(individual))
                sh.cell(row_i, 1).font = Font(bold=True)
                sh.cell(row_i, 1).fill = PatternFill("solid", fgColor="D9EAF7")
                sh.cell(row_i, 2, f"{group['recording'].nunique()} recordings")
                sh.cell(row_i, 2).font = Font(italic=True, color="666666")
                for col in range(1, len(headers) + 1):
                    sh.cell(row_i, col).border = Border(bottom=thin)
                row_i += 1

                # Recording rows under the individual.
                for _, rec in group.iterrows():
                    sh.cell(row_i, 1, str(rec["recording"]))
                    sh.cell(row_i, 2, str(rec["label_name"]))
                    sh.cell(row_i, 3, str(rec["image_name"]))
                    sh.cell(row_i, 4, str(rec["sample_dir"]))
                    sh.cell(row_i, 5, str(rec["hypergui_dir"]))
                    sh.cell(row_i, 6, str(rec["spectrum_xlsx_path"]))
                    for col in range(1, len(headers) + 1):
                        sh.cell(row_i, col).border = Border(bottom=hair)
                        sh.cell(row_i, col).alignment = Alignment(wrap_text=False, vertical="top")
                    row_i += 1

        sh.freeze_panes = "A4"
        sh.auto_filter.ref = f"A3:F{max(row_i - 1, 3)}"
        widths = {"A": 38, "B": 16, "C": 52, "D": 70, "E": 48, "F": 70}
        for col, width in widths.items():
            sh.column_dimensions[col].width = width
        for row in range(4, sh.max_row + 1):
            sh.row_dimensions[row].height = 15


    # Matrix sheets
    for sheet_name, matrix, fmt in [("Matrix Normalized", mat_norm, "0.0000"), ("Matrix Counts", mat_counts, "0")]:
        sh = wb.create_sheet(sheet_name)
        df_mat = pd.DataFrame(matrix, index=labels, columns=labels).reset_index().rename(columns={"index": "true_label"})
        write_df(sh, df_mat)
        sh.column_dimensions["A"].width = 18
        for col in range(2, sh.max_column + 1):
            sh.column_dimensions[get_column_letter(col)].width = 11
            sh.cell(1, col).alignment = Alignment(text_rotation=45, horizontal="center")
        for row in range(2, sh.max_row + 1):
            for col in range(2, sh.max_column + 1):
                sh.cell(row, col).number_format = fmt
                sh.cell(row, col).alignment = Alignment(horizontal="center")
        rng = f"B2:{get_column_letter(sh.max_column)}{sh.max_row}"
        if fmt in ("0.00", "0.0000"):
            sh.conditional_formatting.add(rng, ColorScaleRule(start_type="num", start_value=0, start_color="440154", mid_type="num", mid_value=0.5, mid_color="21908C", end_type="num", end_value=1, end_color="FDE725"))
        else:
            sh.conditional_formatting.add(rng, ColorScaleRule(start_type="min", start_color="FFFFFF", end_type="max", end_color="63BE7B"))

    # Settings sheets
    settings_rows = flatten_dict(yaml_settings) if isinstance(yaml_settings, dict) else []
    settings_df = pd.DataFrame(settings_rows)
    compact = settings_df[~settings_df["setting"].astype(str).str.startswith(("experiment_folders.", "labelling_file.", "hyperguis.", "paths."))].copy() if not settings_df.empty else pd.DataFrame(columns=["setting", "value"])
    meta_df = pd.DataFrame([{"setting": f"metadata.{k}", "value": v} for k, v in metadata.items()])
    compact = pd.concat([meta_df, compact], ignore_index=True)
    sh = wb.create_sheet("Settings Summary")
    write_df(sh, compact)
    sh.column_dimensions["A"].width = 44
    sh.column_dimensions["B"].width = 42

    full_parts = []
    for section, data in [("yaml_settings", yaml_settings), ("run_summary", run_summary), ("generated_config", generated_configs)]:
        rows = flatten_dict(data) if isinstance(data, dict) else []
        if rows:
            df = pd.DataFrame(rows)
            df.insert(0, "section", section)
            full_parts.append(df)
    full = pd.concat(full_parts, ignore_index=True) if full_parts else pd.DataFrame(columns=["section", "setting", "value"])
    sh = wb.create_sheet("Full Settings and Paths")
    write_df(sh, full)
    sh.column_dimensions["A"].width = 20
    sh.column_dimensions["B"].width = 55
    sh.column_dimensions["C"].width = 90
    for row in range(2, sh.max_row + 1):
        sh.row_dimensions[row].height = 15
        sh.cell(row, 3).alignment = Alignment(wrap_text=False)

    for sh in wb.worksheets:
        sh.sheet_view.showGridLines = False

    wb.save(xlsx_path)


def compute_split_counts(output_dir: Path, labels: list[str]) -> tuple[pd.DataFrame, dict[str, Any]]:
    manifest_path = output_dir / "median_pixel_yaml_manifest.csv"
    data_spec_path = output_dir / "yaml_generated_median_pixel_data_spec_htc_adapter_annotated.json"
    info: dict[str, Any] = {"manifest_path": str(manifest_path), "data_spec_path": str(data_spec_path), "method": "manifest_vs_data_spec_key_matching"}

    if not manifest_path.exists() or not data_spec_path.exists():
        return pd.DataFrame(columns=["split", "label", "samples"]), {**info, "warning": "manifest or data_spec missing"}

    manifest = pd.read_csv(manifest_path)
    spec = json.loads(data_spec_path.read_text(encoding="utf-8"))[0]

    name_col = "htc_image_name_annotation" if "htc_image_name_annotation" in manifest.columns else "image_name"
    if name_col not in manifest.columns or "label_name" not in manifest.columns:
        return pd.DataFrame(columns=["split", "label", "samples"]), {**info, "warning": "required manifest columns missing"}

    rows = []
    unmatched = {}
    for split in ["train", "val", "test"]:
        split_names = spec.get(split, {}).get("image_names", [])
        spec_keys = set()
        for name in split_names:
            spec_keys |= row_keys(str(name))

        mask = manifest[name_col].astype(str).apply(lambda x: bool(row_keys(x) & spec_keys))
        d = manifest[mask]
        counts = d["label_name"].astype(str).value_counts().to_dict()
        for label in labels:
            rows.append({"split": split, "label": label, "samples": int(counts.get(label, 0))})
        unmatched[split] = {"spec_names": len(split_names), "matched_manifest_rows": int(len(d))}

    table = pd.DataFrame(rows)
    pivot = table.pivot(index="label", columns="split", values="samples").reset_index()
    for col in ["train", "val", "test"]:
        if col not in pivot.columns:
            pivot[col] = 0
    pivot["total"] = pivot[["train", "val", "test"]].sum(axis=1)
    pivot = pivot[["label", "train", "val", "test", "total"]]
    return pivot, {**info, "match_summary": unmatched}


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--run-dir", required=True)
    ap.add_argument("--label-mapping", required=True)
    ap.add_argument("--scenario-dir", required=True)
    ap.add_argument("--output-dir", required=True)
    ap.add_argument("--profile", required=True)
    ap.add_argument("--yaml-path", default="")
    ap.add_argument("--normalize", action="store_true")
    ap.add_argument("--write-sample-predictions", action="store_true", help="Also write long row-level prediction CSV for debugging")
    args = ap.parse_args()

    run_dir = Path(args.run_dir)
    scenario_dir = Path(args.scenario_dir)
    output_dir = Path(args.output_dir)
    yaml_path = Path(args.yaml_path) if args.yaml_path else None
    label_mapping_path = Path(args.label_mapping)
    label_mapping = json.loads(label_mapping_path.read_text(encoding="utf-8"))
    inv = {int(v): str(k) for k, v in label_mapping.items()}
    labels = [k for k, _ in sorted(label_mapping.items(), key=lambda kv: int(kv[1]))]

    pred_path = find_prediction_table(run_dir)
    df = pd.read_csv(pred_path)
    true_col = "label_index_mapped"
    pred_col = "prediction"
    if true_col not in df.columns or pred_col not in df.columns:
        raise ValueError(f"Prediction table missing {true_col}/{pred_col}: {pred_path}")

    df["true_label"] = df[true_col].map(inv)
    df["pred_label"] = df[pred_col].map(inv)
    df["correct"] = df[true_col].eq(df[pred_col])

    n = len(labels)
    mat_counts = np.zeros((n, n), dtype=int)
    for t, p in zip(df[true_col], df[pred_col]):
        if int(t) in inv and int(p) in inv:
            mat_counts[int(t), int(p)] += 1
    row_sums = mat_counts.sum(axis=1, keepdims=True)
    mat_norm = np.divide(mat_counts, row_sums, out=np.zeros_like(mat_counts, dtype=float), where=row_sums != 0)
    mat_to_plot = mat_norm if args.normalize else mat_counts

    acc = float(df["correct"].mean()) if len(df) else float("nan")
    per_class = []
    for i, label in enumerate(labels):
        support = int(mat_counts[i].sum())
        correct = int(mat_counts[i, i])
        sensitivity = correct / support if support else np.nan
        pred_count = int(mat_counts[:, i].sum())
        per_class.append({"label": label, "test_support": support, "predicted_count": pred_count, "correct": correct, "sensitivity": sensitivity})

    scenario_dir.mkdir(parents=True, exist_ok=True)
    prefix = scenario_dir / f"confusion_matrix_{args.profile}"
    xlsx_path = prefix.with_name(prefix.name + "_report.xlsx")

    yaml_settings = {}
    if yaml_path and yaml_path.exists():
        yaml_settings = yaml.safe_load(yaml_path.read_text(encoding="utf-8")) or {}
    run_summary = safe_load_json(output_dir / "run_summary.json")
    generated_configs = {p.name: safe_load_json(p) for p in output_dir.glob("generated_median_pixel_config*.json")}
    report_metadata = {
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "accuracy": acc,
        "n_test_rows": len(df),
        "profile": args.profile,
        "run_dir": str(run_dir),
        "prediction_table": str(pred_path),
        "label_mapping": str(label_mapping_path),
        "yaml_path": str(yaml_path) if yaml_path else "",
    }

    split_counts, split_info = compute_split_counts(output_dir, labels)
    individual_counts, recording_counts, split_details, split_assignment_info = build_split_assignment(output_dir, labels)

    # CSV 1: compact metrics + matrix + settings.
    metrics_csv = prefix.with_name(prefix.name + "_metrics_matrix.csv")
    with metrics_csv.open("w", encoding="utf-8", newline="") as f:
        f.write("# report_type,confusion_metrics_matrix\n")
        f.write(f"# created_at,{datetime.now().isoformat(timespec='seconds')}\n")
        f.write(f"# accuracy,{acc}\n")
        f.write(f"# n_test_rows,{len(df)}\n")
        f.write(f"# profile,{args.profile}\n")
        f.write(f"# run_dir,{run_dir}\n")
        f.write(f"# prediction_table,{pred_path}\n")
        f.write(f"# label_mapping,{label_mapping_path}\n")
        if yaml_path:
            f.write(f"# yaml_path,{yaml_path}\n")

        f.write("\n[yaml_settings]\n")
        pd.DataFrame(flatten_dict(yaml_settings)).to_csv(f, index=False)

        f.write("\n[run_summary]\n")
        pd.DataFrame(flatten_dict(run_summary)).to_csv(f, index=False)

        f.write("\n[generated_config]\n")
        pd.DataFrame(flatten_dict(generated_configs)).to_csv(f, index=False)

        f.write("\n[per_class_metrics]\n")
        pd.DataFrame(per_class).to_csv(f, index=False)

        f.write("\n[confusion_matrix_normalized]\n")
        pd.DataFrame(mat_norm, index=labels, columns=labels).to_csv(f, index_label="true_label")

        f.write("\n[confusion_matrix_counts]\n")
        pd.DataFrame(mat_counts, index=labels, columns=labels).to_csv(f, index_label="true_label")

    # CSV 2: split sample counts, not long sample-level prediction rows.
    split_csv = prefix.with_name(prefix.name + "_split_samples.csv")
    with split_csv.open("w", encoding="utf-8", newline="") as f:
        f.write("# report_type,split_sample_counts\n")
        f.write(f"# profile,{args.profile}\n")
        f.write(f"# run_dir,{run_dir}\n")
        f.write(f"# data_spec,{split_info.get('data_spec_path','')}\n")
        f.write(f"# manifest,{split_info.get('manifest_path','')}\n")
        f.write("\n[split_counts]\n")
        split_counts.to_csv(f, index=False)

    # Optional long debug CSV.
    sample_pred_csv = None
    if args.write_sample_predictions:
        sample_pred_csv = prefix.with_name(prefix.name + "_sample_predictions_debug.csv")
        sample_cols = [c for c in ["image_name", "image_name_annotations", "image_index", "batch_idx", "row_in_batch"] if c in df.columns]
        samples = df[sample_cols + [true_col, pred_col, "true_label", "pred_label", "correct"]].copy()
        samples.insert(0, "profile", args.profile)
        samples.insert(1, "run_dir", str(run_dir))
        samples.insert(2, "prediction_table", str(pred_path))
        samples.to_csv(sample_pred_csv, index=False)

    # PNG + HTML report.
    png_path = prefix.with_suffix(".png")
    pdf_path = prefix.with_suffix(".pdf")
    fig, ax = plt.subplots(figsize=(max(8, n * 0.7), max(6, n * 0.55)))
    im = ax.imshow(mat_to_plot, aspect="auto")
    ax.set_xticks(np.arange(n)); ax.set_yticks(np.arange(n))
    ax.set_xticklabels(labels, rotation=45, ha="right")
    ax.set_yticklabels(labels)
    ax.set_xlabel("Predicted")
    ax.set_ylabel("True")
    ax.set_title(f"Confusion matrix normalized | accuracy={acc:.4f}")
    for i in range(n):
        for j in range(n):
            value = mat_to_plot[i, j]
            if value != 0:
                ax.text(j, i, f"{value:.2f}" if args.normalize else str(int(value)), ha="center", va="center", color="black", fontsize=8)
    fig.colorbar(im, ax=ax)
    fig.tight_layout()
    fig.savefig(png_path, dpi=200)
    fig.savefig(pdf_path)
    plt.close(fig)

    html_path = prefix.with_suffix(".html")
    img_b64 = base64.b64encode(png_path.read_bytes()).decode("ascii")
    per_class_html = pd.DataFrame(per_class).to_html(index=False, float_format=lambda x: f"{x:.4f}")
    split_html = split_counts.to_html(index=False) if not split_counts.empty else "<p>No split counts available.</p>"
    settings_html = pd.DataFrame(flatten_dict(yaml_settings)).to_html(index=False) if yaml_settings else "<p>No YAML settings available.</p>"
    html = f"""<!doctype html>
<html><head><meta charset='utf-8'><title>Confusion Report {args.profile}</title>
<style>body{{font-family:Arial,sans-serif;margin:24px}} code,pre{{background:#f5f5f5;padding:4px}} table{{border-collapse:collapse;margin-bottom:24px}} td,th{{border:1px solid #ccc;padding:4px 8px}}</style>
</head><body>
<h1>Confusion Matrix Report: {args.profile}</h1>
<p><b>Accuracy:</b> {acc:.4f}<br><b>Test rows:</b> {len(df)}<br><b>Run dir:</b> <code>{run_dir}</code><br><b>Prediction table:</b> <code>{pred_path}</code></p>
<h2>Confusion Matrix</h2><img src='data:image/png;base64,{img_b64}' style='max-width:100%;height:auto'>
<h2>Per-class test metrics</h2>{per_class_html}
<h2>Train/validation/test sample counts</h2>{split_html}
<h2>YAML settings</h2>{settings_html}
<h2>Output files</h2><ul><li>{png_path}</li><li>{xlsx_path}</li><li>{pdf_path}</li><li>{xlsx_path}</li>{'<li>'+str(sample_pred_csv)+'</li>' if sample_pred_csv else ''}</ul>
</body></html>"""
    html_path.write_text(html, encoding="utf-8")

    write_clean_excel_report(
        xlsx_path=xlsx_path,
        png_path=png_path,
        labels=labels,
        per_class=per_class,
        mat_norm=mat_norm,
        mat_counts=mat_counts,
        individual_counts=individual_counts,
        recording_counts=recording_counts,
        split_details=split_details,
        yaml_settings=yaml_settings,
        run_summary=run_summary,
        generated_configs=generated_configs,
        metadata=report_metadata,
    )

    # The XLSX workbook now contains the metrics, matrices, split counts, and split details.
    # Remove the auxiliary CSV files to keep the scenario folder clean.
    for _csv_path in [metrics_csv, split_csv]:
        try:
            Path(_csv_path).unlink(missing_ok=True)
        except Exception as _e:
            print(f"[WARNING] Could not remove auxiliary CSV {_csv_path}: {_e}")

    print(f"[OK] accuracy: {acc:.4f}")
    print("[OK] auxiliary CSV outputs removed; XLSX contains spreadsheet report")
    if sample_pred_csv:
        print(f"[OK] wrote: {sample_pred_csv}")
    print(f"[OK] wrote: {png_path}")
    print(f"[OK] wrote: {pdf_path}")
    print(f"[OK] wrote: {xlsx_path}")
    print(f"[OK] wrote: {html_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
